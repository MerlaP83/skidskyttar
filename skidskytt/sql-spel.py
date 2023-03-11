import random
import time
import sys,os
from colorama import init as colorama_init
from colorama import Fore
from colorama import Style
from prettytable import PrettyTable
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import sqlite3

# Connect to the SQLite database
con = sqlite3.connect('skidskytte.db')

# Prompt the user to enter their choice of "damer" or "herrar"
val = input("Välj damer (1) eller herrar (2): ")

# Define the SQL query based on user input
if val == "1":
    gender = "f"
    sql = 'SELECT * FROM skidskyttar WHERE kon = ?'
elif val == "2":
    gender = "m"
    sql = 'SELECT * FROM skidskyttar WHERE kon = ?'
else:
    print("Felaktigt val. Avslutar programmet.")
    sys.exit()

# Hämta tävlingsdata från tabellen "competitions"
con.row_factory = sqlite3.Row
cursor = con.cursor()

# Hämta tävlingar som inte finns i resultat-tabellen
cursor.execute("SELECT * FROM tavling WHERE kon = ? ORDER BY datum ASC", (gender,))
tavlingar = cursor.fetchall()

# Låt användaren välja tävling
print("Välj tävling:")
for i, tavling in enumerate(tavlingar):
    # Kontrollera om tävlingen redan finns i resultat-tabellen
    cursor.execute("SELECT COUNT(*) FROM resultat WHERE tavling = ? AND gren = ? AND kon = ?", (tavling['namn'], tavling['gren'], tavling['kon'],))
    if cursor.fetchone()[0] > 0:
        continue  # Tävlingen är redan avgjord, fortsätt till nästa
    print(f"{i+1}. {tavling['datum']}: {tavling['namn']} - {tavling['gren']}.")
try:
    vald_tavling_index = int(input("Ange vald tävling: ")) - 1
    vald_tavling = tavlingar[vald_tavling_index]
except (ValueError, IndexError):
    print("Felaktig inmatning, väljer första tillgängliga tävlingen som default.")
    vald_tavling = tavlingar[0]


if vald_tavling['gren'] == 'Jaktstart':
    # Hämta de 40 åkarna med lägst tidstillägg och starttid > 0
    cursor.execute(f"SELECT * FROM skidskyttar WHERE kon = ? AND starttid > 0 ORDER BY tidstillagg ASC LIMIT 40", (gender,))
    skidskyttar = []
    weights = [8, 1, 1]  # Viktning för varje alternativ: [0, -1, 1]
    for row in cursor:
        skidskytt = dict(row)
        skidskytt["form"] = random.choices([0, -1, 1], weights=weights)[0]
        skidskyttar.append(skidskytt)

else:
    # Hämta de 40 åkarna med lägst tidstillägg
    cursor.execute(f"SELECT * FROM skidskyttar WHERE kon = ? ORDER BY tidstillagg ASC LIMIT 30", (gender,))
    skidskyttar = []
    weights = [8, 1, 1]  # Viktning för varje alternativ: [0, -1, 1]
    for row in cursor:
        skidskytt = dict(row)
        skidskytt["form"] = random.choices([0, -1, 1], weights=weights)[0]
        skidskyttar.append(skidskytt)

    # Välj ytterligare 10 slumpmässiga åkare från databasen som inte redan finns i skidskyttar
    cursor.execute(sql, (gender,))
    remaining_skidskyttar = [dict(row) for row in cursor if dict(row)["id"] not in [sk["id"] for sk in skidskyttar]]
    additional_skidskyttar = random.sample(remaining_skidskyttar, 10)

    # Lägg till de slumpmässiga åkarna till listan skidskyttar
    skidskyttar += additional_skidskyttar


# Grundtid per kilometer i sekunder
grundtid = 130

bra_form = []
dalig_form = []

# Beräkningar för förhandsfavoriter
for skidskytt in skidskyttar:
    fav_tidstillagg = skidskytt["tidstillagg"] + skidskytt["form"]
    fav_form = skidskytt["form"]
    fav_liggande_traffsakerhet = skidskytt["liggande_traffsakerhet"] / 100
    fav_staende_traffsakerhet = skidskytt["staende_traffsakerhet"] / 100
    fav_viktad_poang = (fav_tidstillagg * 3) + 100 / (fav_liggande_traffsakerhet + fav_staende_traffsakerhet) + fav_form
    skidskytt["viktad_poäng"] = fav_viktad_poang + random.randint(-3, 3)

# Sortera förhandsfavoriterna efter viktad poäng och välj de fem bästa
favoriter = sorted(skidskyttar, key=lambda x: x["viktad_poäng"])[:5]

def main_menu():
    def run_menu():
        clear_screen()
        print("SKIDSKYTTE SPEL")
        print("===============")
        print("1. Ny tävling")
        print("2. Se resultat")
        print("3. Se skidskyttar")
        print("4. Avsluta")
        choice = input("Välj ett alternativ (1-4): ")

        if choice == "1":
            os.system("sql-spel.py")
        elif choice == "2":
            os.system("resultat.py")
        elif choice == "3":
            os.system("skidskyttar.py")
        elif choice == "4":
            clear_screen()
            print("Tack för att du spelade! Hejdå!")
            return False
        else:
            print("Ogiltigt val. Försök igen.")
        return True

    while run_menu():
        pass

# Funktion för att skriva ut resultatlistan och spara den i databasen
def print_stallning(skidskyttar, tavling):
    print(f"\nResultat - {tavling['namn']} ({tavling['gren']}, {tavling['distans']} km):")
    skidskyttar = sorted(skidskyttar, key=lambda x: x["total_tid"])
    medaljer = ["🥇 ", "🥈 ", "🥉 "]
    with sqlite3.connect('skidskytte.db') as con:
        cursor = con.cursor()
        # Get the latest competition ID from the "resultat" table
        cursor.execute("SELECT MAX(id) FROM resultat")
        latest_id = cursor.fetchone()[0]
        # Calculate the new competition ID (or set to 1 if there are no previous results)
        if latest_id is None:
            new_id = 1
        else:
            new_id = latest_id + 1
        for i, skidskytt in enumerate(skidskyttar):
            tid_sedan_ledare = round(skidskytt["total_tid"] - skidskyttar[0]["total_tid"], 1)
            tid_sedan_ledare_min, tid_sedan_ledare_sec = divmod(int(tid_sedan_ledare), 60)
            total_min, total_sec = divmod(int(skidskytt["total_tid"]), 60)
            total_tid_str = f"{total_min:02d}:{total_sec:02d}"
            tid_sedan_ledare_str = f"+{tid_sedan_ledare_min}m {tid_sedan_ledare_sec}s" if tid_sedan_ledare > 0 else ""
            if i < 3:
                prefix = medaljer[i]
                print(f"{prefix}{skidskytt['namn']} ({total_tid_str} {tid_sedan_ledare_str})"
                      f" - straffrundor: {skidskytt['straffrundor_totalt']}.")
            else:
                print(f"{i+1}. {skidskytt['namn']} ({total_tid_str} {tid_sedan_ledare_str})"
                      f" - straffrundor: {skidskytt['straffrundor_totalt']}.")

            if vald_tavling['typ'] == 'Världscupen':
                if i < 41:
                    poang_lista = [90, 75, 60, 50, 45, 40, 36, 34, 32, 31, 30, 29, 28, 27, 26, 25, 24, 23, 22, 21, 20,
                                   19, 18, 17, 16, 15, 14, 13, 12, 11, 10, 9, 8, 7, 6, 5, 4, 3, 2, 1]
                    vpoang = poang_lista[i]
                    cursor.execute(
                        "UPDATE skidskyttar SET vpoang = vpoang + ? WHERE id = ?",
                        (vpoang, skidskytt["id"])
                    )

            if tavling["gren"] == "Sprint":
                starttid = tid_sedan_ledare + 0.1
                cursor.execute(
                    "UPDATE skidskyttar SET starttid = ? WHERE id = ?",
                    (starttid, skidskytt["id"])
                )
            else:  # Lägg till en else-sats och tilldela starttid till 0.
                starttid = 0

            if tavling["gren"] == "Jaktstart":
                cursor.execute(
                    "UPDATE skidskyttar SET starttid = 0 WHERE id = ?",
                    (skidskytt["id"],)
                )

            cursor.execute(
                "INSERT INTO resultat (placering, akarid, namn, total_tid, tid_sedan_ledare, straffrundor_totalt, id, tavling, gren, typ, kon) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (i + 1, skidskytt["id"], skidskytt["namn"], skidskytt["total_tid"], tid_sedan_ledare_str,
                 skidskytt["straffrundor_totalt"], new_id, tavling["namn"], tavling["gren"], tavling["typ"],tavling["kon"])
            )

    # Skriv ut topp 5 i världscupen
    topp_5 = [skidskytt for skidskytt in skidskyttar
              if skidskytt['vpoang'] > 0 and skidskytt['kon'] == vald_tavling['kon']]
    topp_5 = sorted(topp_5, key=lambda x: x['vpoang'], reverse=True)[:5]

    if topp_5:
        t = PrettyTable(['Namn', 'Nation', 'Världscuppoäng'])
        for skidskytt in topp_5:
            t.add_row([skidskytt['namn'], skidskytt['nation'], f"{skidskytt['vpoang']:.2f}"])
        print(Fore.GREEN + "Topp 5 i världscupen:" + Style.RESET_ALL)
        print(t)
    else:
        print(Fore.GREEN + "Topp 5 i världscupen:" + Style.RESET_ALL)
        print("Det här är första världscuptävlingen.")


# Slumpmässiga händelser

def hantera_handelse(skidskytt, tid, kilometer):
    colorama_init()

    handelser = {
        1: {"tid": 10, "text": "får problem med ena skidan!"},
        2: {"tid": 15, "text": "ena stav går av!"},
        3: {"tid": 5, "text": "verkar ha dåligt glid!"},
        4: {"tid": 10, "text": "vurpar!"},
        5: {"tid": 5, "text": "tappar en stav och tvingas hämta den."},
        6: {"tid": -5, "text": "får en burst av energi och ökar tempot."},
        7: {"tid": -3, "text": "kämpar på bra."},
        8: {"tid": -5, "text": "kämpar extra hårt efter publikens hejarop."},
        9: {"tid": -3, "text": "åker tekniskt skickligt idag."},
        10: {"tid": -8, "text": "åker överraskande snabbt!"},
        11: {"tid": -5, "text": "har hittat ett bättre glid och ökar tempot."},
        12: {"tid": -5, "text": "ökar takten och ser ut att orka hålla ett högre tempo."},
        13: {"tid": -5, "text": "är stark i motlut och tar in tid på konkurrenterna."},
        14: {"tid": -3, "text": "är en av de snabbaste i nedförsbackarna."},
        15: {"tid": 10, "text": "tappar en skida och tvingas stanna för att sätta fast den igen."},
        16: {"tid": 15, "text": "kraschar in i en annan åkare och tappar tid."},
        17: {"tid": 8, "text": "känner av en muskelskada och tvingas sänka tempot."},
        18: {"tid": 15, "text": "åker fel och måste vända om för att komma tillbaka till rätt bana."},
        19: {"tid": 5, "text": "ser tung ut i åkningen."},
        20: {"tid": 10, "text": "har problem med en skidbindning och tvingas stanna för att åtgärda det."},
        21: {"tid": 15, "text": "ser ut att ha vallat bort sig idag."},
        22: {"tid": -5, "text": "har bra flyt i åkningen."},
        23: {"tid": -5, "text": "får extra energi och orkar åka fortare."},
        24: {"tid": -7, "text": "har utvecklats i tekniken och åker betydligt bättre idag än tidigare."},
        25: {"tid": -8, "text": "har en extra stark dag och kan åka med mycket kraft i åkningen."},
        26: {"tid": -10, "text": "har bättre glid än konkurrenterna!"},
        27: {"tid": 20, "text": "har helt kroknat och ser inte frisk ut idag. Aj, aj, aj!"}
    }

    handelse = random.randint(1, 600)
    if handelse in handelser:
        handelsebeskrivning = handelser.get(handelse)
        tid += handelsebeskrivning["tid"]
        skidskytt[f"kilometer_{kilometer}_handelse"] = handelsebeskrivning["text"]

        # Byt textfärg beroende på om tiden är negativ eller positiv
        textfarg = Fore.GREEN if handelsebeskrivning["tid"] < -0.00001 else Fore.RED
        textfarg = Fore.WHITE if abs(handelsebeskrivning["tid"]) <= 0.00001 else textfarg
        print(f"{textfarg}{skidskytt['namn']}{Style.RESET_ALL} {handelsebeskrivning['text']}")

        time.sleep(1)

    # Beräkna total tid för varje skidskytt för varje kilometer och totalt
    skidskytt[f"kilometer_{kilometer}_total_tid"] = (skidskytt.get(f"kilometer_{kilometer-1}_total_tid") or 0)
    skidskytt["total_tid"] = skidskytt[f"kilometer_{kilometer}_total_tid"]

    return tid

# Funktion för att uppdatera träffsäkerheten beroende på vinden
def uppdatera_traffsakerhet(skidskytt, skyttelage):
    skyttelag = skidskytt.get(f"{skyttelage}_traffsakerhet")
    if skyttelag is None:
        print("Felaktigt skyttelage angivet.")
        return

    vindstyrka = vindstyrka1 + random.randint(-2, 2)
    beskrivningar = {-6: "Mycket bra", -5: "Bra", -4: "Bra", -3: "Bra", -2: "Bra", -1: "Bra", 0: "Neutrala", 1: "Svåra",
                     2: "Svåra", 3: "Svåra", 4: "Mycket svåra", 5: "Mycket svåra", 6: "Extremt svåra"}
    beskrivning = beskrivningar.get(vindstyrka, "Ogiltig vindstyrka") + " vindförhållanden."

    print("Aktuell vindstyrka:",beskrivning)

    ny_traffsakerhet = skyttelag - vindstyrka - (skidskytt["form"] * 5)
    skidskytt[f"{skyttelage}_traffsakerhet"] = max(min(ny_traffsakerhet, 100), 0)

def save_excel(skidskyttar, tavling):
    # Skapa rubriker
    headers = [
        "#",
        "Namn",
        "Tid",
        "Straffrundor",
        "Li-T",
        "Li-S",
        "Li-%",
        "St-T",
        "St-S",
        "St-%",
        "VC-poäng"
    ]

    # Sortera skidskyttar baserat på totala tiden och lagra placeringen
    skidskyttar = sorted(skidskyttar, key=lambda k: k['total_tid'])
    for i, skidskytt in enumerate(skidskyttar):
        skidskytt['placering'] = i + 1

    # Skapa en lista med data för varje skidskytt
    skidskytt_data = [
        {
            "#": skidskytt["placering"],
            "Namn": skidskytt["namn"],
            "Tid": f'{int(skidskytt["total_tid"] // 60)}:{int(skidskytt["total_tid"] % 60):02}',
            "Straffrundor": skidskytt["straffrundor_totalt"],
            "Li-T": skidskytt["liggande_traffar"],
            "Li-S": 10 if len(liggande_skytte_index) == 2 else 5 if len(liggande_skytte_index) == 1 else 0,
            "St-T": skidskytt["staende_traffar"],
            "St-S": 10 if len(staende_skytte_index) == 2 else 5 if len(staende_skytte_index) == 1 else 0,
            "VC-poäng": skidskytt["vpoang"]
        }
        for skidskytt in skidskyttar
    ]

    # Uppdatera träffprocenten efter att skidskytt_data har skapats
    for i, data in enumerate(skidskytt_data):
        skidskytt_data[i][
            "Li-%"] = f'{(skidskytt_data[i]["Li-T"] / int(skidskytt_data[i]["Li-S"])) * 100:.0f}'
        skidskytt_data[i][
            "St-%"] = f'{(skidskytt_data[i]["St-T"] / int(skidskytt_data[i]["St-S"])) * 100:.0f}'

    # Skapa ett Excel-dokument
    workbook = Workbook()
    worksheet = workbook.active

    # Skriv rubrikerna till Excel
    for i, header in enumerate(headers):
        # Lägg till information om tävlingen
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        worksheet.cell(row=1, column=1,
                       value=f"Tävling: {tavling['namn']} ({tavling['gren']}, {tavling['distans']} km)").font = Font(
            bold=True)
        worksheet.cell(row=1, column=8, value="Datum").font = Font(bold=True)
        worksheet.cell(row=1, column=9, value=tavling['datum']).number_format = 'yyyy-mm-dd'
        worksheet.cell(row=3, column=i+1, value=header).font = Font(bold=True)
        worksheet.column_dimensions[get_column_letter(i+1)].width = len(header) + 2
        worksheet.column_dimensions[get_column_letter(i+1)].bestFit = True # Automatisk justering av kolumnbredd

        # Lägg till Autofilter
        worksheet.auto_filter.ref = "A3:K3"

    # Skriv data till Excel
    for i, data in enumerate(skidskytt_data):
        for j, header in enumerate(headers):
            worksheet.cell(row=i+4, column=j+1, value=data[header]).alignment = Alignment(horizontal="center")

    # Spara Excel-dokumentet
    filename = f"{tavling['namn']}_{tavling['datum']}_{tavling['kon']}.xlsx"
    workbook.save(filename)

    print(f"Resultaten har sparats till filen {filename}")


def clear_screen():
    os.system("cls" if os.name == "nt" else "clear")


# Definiera antal kilometer
antal_kilometer = vald_tavling['distans']

# Definiera index för kilometer med liggande skytte
liggande_skytte_index = vald_tavling['liggande_skytte']
if isinstance(liggande_skytte_index, int):
    liggande_skytte_index = [liggande_skytte_index]
else:
    liggande_skytte_index = [int(x) for x in liggande_skytte_index.split(',')]
staende_skytte_index = vald_tavling['staende_skytte']
if isinstance(staende_skytte_index, int):
    staende_skytte_index = [staende_skytte_index]
else:
    staende_skytte_index = [int(x) for x in staende_skytte_index.split(',')]


# Definiera ett högt startvärde för ledande tid
ledande_tid_min = 0
ledande_tid_sec = 0
ledande_skytt = ""

# Slumpa fram vindstyrkan
vindstyrka1 = random.randint(-1, 1)

# Spelet börjar
print(f"\nVälkomna till {vald_tavling['namn']} i {vald_tavling['plats']}.\nI dag blir det en {vald_tavling['gren']} på"
      f" {vald_tavling['distans']} km. Det kommer att bli liggande skytte på kilometer "
      f"{vald_tavling['liggande_skytte']} och stående skytte på kilometer {vald_tavling['staende_skytte']} .\n")

# initiera colorama
colorama_init(autoreset=True)

# Skriv ut favoriterna
t = PrettyTable(['Namn', 'Nation', 'Världscuppoäng'])
for skidskytt in favoriter:
    t.add_row([skidskytt['namn'], skidskytt['nation'], skidskytt['vpoang']])
print(Fore.GREEN + "Förhandsfavoriter:" + Style.RESET_ALL)
print(t)

time.sleep(2)

# Skriv ut topp 5 i världscupen
topp_5 = [skidskytt for skidskytt in skidskyttar
          if skidskytt['vpoang'] > 0 and skidskytt['kon'] == vald_tavling['kon']]
topp_5 = sorted(topp_5, key=lambda x: x['vpoang'], reverse=True)[:5]

if topp_5:
    t = PrettyTable(['Namn', 'Nation', 'Världscuppoäng'])
    for skidskytt in topp_5:
        t.add_row([skidskytt['namn'], skidskytt['nation'], f"{skidskytt['vpoang']:.2f}"])
    print(Fore.GREEN + "Topp 5 i världscupen:" + Style.RESET_ALL)
    print(t)
else:
    print(Fore.GREEN + "Topp 5 i världscupen:" + Style.RESET_ALL)
    print("Det här är första världscuptävlingen.")


time.sleep(2)

# Skriv ut namnen på skidskyttarna med bra form
bra_form = [skidskytt["namn"] for skidskytt in skidskyttar if skidskytt["form"] == -1]
dalig_form = [skidskytt["namn"] for skidskytt in skidskyttar if skidskytt["form"] == 1]
print(f"\nSkidskyttar som väntas ha en bra dag är: {', '.join([Fore.GREEN + skidskytt + Style.RESET_ALL for skidskytt in bra_form])}. "
      f"\nDe som väntas ha en sämre dag är: {', '.join([Fore.RED + skidskytt + Style.RESET_ALL for skidskytt in dalig_form])}.\n")

if vindstyrka1 <= -1:
    print("Bra vindförhållanden.")
elif vindstyrka1 <= 0:
    print("Måttliga vindförhållanden.")
else:
    print("Svåra vindförhållanden.")

time.sleep(4)

for kilometer in range(antal_kilometer):
    # Sortera skidskyttarna efter total tid för den aktuella kilometer
    stallning = sorted(skidskyttar, key=lambda x: x.get(f"kilometer_{kilometer}_total_tid", float('inf')))
    plats = 1
    for skidskytt in stallning:
        skidskytt[f"kilometer_{kilometer}_position"] = plats
        plats += 1
    print(f"\nKilometer {kilometer + 1} av {antal_kilometer}:")

    # Sortera skidskyttarna efter total tid över alla kilometer
    stallning = sorted(skidskyttar, key=lambda x: x["total_tid"])
    ledare_tid = stallning[0]["total_tid"]

    for i, skidskytt in enumerate(stallning):
        form = skidskytt["form"]
        tarning = random.randint(1, 6)
        tid = form + tarning
        if kilometer == 0 and vald_tavling['gren'] == 'Jaktstart':
            tid += skidskytt['starttid']

        if kilometer not in liggande_skytte_index and kilometer not in staende_skytte_index:
            tid = hantera_handelse(skidskytt, tid, kilometer)

        skidskytt[f"kilometer_{kilometer}_tid"] = tid

        # Liggande skytte
        if kilometer in liggande_skytte_index:

            # Uppdatera skyttens träffsäkerhet
            skyttelage = "liggande"

            # Rätt ordning vid skidskyttet

            index = stallning.index(skidskytt)
            if index < len(stallning) - 1:
                stallning.insert(index + 1, stallning.pop(index + 1))

            total_min, total_sec = divmod(int(skidskytt["total_tid"]), 60)
            tid_sedan_ledare = ledare_tid - skidskytt["total_tid"]
            tid_sedan_ledare_min = -int(tid_sedan_ledare) // 60
            tid_sedan_ledare_sec = abs(int(tid_sedan_ledare)) % 60

            if tid_sedan_ledare == 0:
                print(
                    f"{i + 1}. {skidskytt['namn']}({skidskytt['straffrundor_totalt']}) - {total_min:02d}:"
                    f"{total_sec:02d}, i ledning. Träffsäkerhet: {skidskytt[skyttelage + '_traffsakerhet']}%.")
            else:
                print(
                    f"{i + 1}. {skidskytt['namn']}({skidskytt['straffrundor_totalt']}) - {total_min:02d}:"
                    f"{total_sec:02d}, {abs(tid_sedan_ledare_min)}m {tid_sedan_ledare_sec}s efter ledaren."
                    f" Träffsäkerhet: {skidskytt[skyttelage + '_traffsakerhet']}%.")

            # Uppdatera träffsäkerheten och utför det liggande skyttet
            ursprunglig_liggande_traffsakerhet = skidskytt["liggande_traffsakerhet"]
            uppdatera_traffsakerhet(skidskytt, "liggande")
            straffrundor = 0
            traffar = 0
            for skott in range(5):
                slump = random.randint(1, 100)
                if slump <= skidskytt[f"{skyttelage}_traffsakerhet"]:
                    traffar += 1
                else:
                    straffrundor += 1
                print(
                    f"Skott {skott + 1}: Träffar: {traffar} av 5 - straffrundor {straffrundor}.")
                # Vänta längre tid vid de första skyttarna för att skapa spänning
                if i < 10:
                    time.sleep(0.8)
                elif i > 15:
                    time.sleep(0.1)
                else:
                    time.sleep(0)

            ursprunglig_liggande_tid = skidskytt["liggande_tid"]
            liggande_tid = skidskytt["liggande_tid"] + random.randint(-5, 5)
            print(f"Liggande tid: {liggande_tid} sekunder.")

            if traffar == 5:
                print(f"Fantastiskt! {Fore.GREEN}{skidskytt['namn']} skjuter fullt!{Style.RESET_ALL}")
            elif traffar == 3:
                print(f"Svagt skytte av {skidskytt['namn']}.")
            elif traffar < 3:
                print(f"{Fore.RED}Vilken mardröm för {skidskytt['namn']}!{Style.RESET_ALL}")

            straff_tidstillagg = 20
            tid = liggande_tid + (straffrundor * straff_tidstillagg)
            skidskytt.update({
                "liggande_tid": ursprunglig_liggande_tid,
                "liggande_traffar": skidskytt.get("liggande_traffar", 0) + traffar,
                "straffrundor": straffrundor,
                "straffrundor_totalt": skidskytt.get("straffrundor_totalt", 0) + straffrundor,
                "liggande_traffsafkerhet": ursprunglig_liggande_traffsakerhet
            })
            print(
                f"Total tid för skyttet: {tid} sekunder ({traffar} träffar, {straffrundor} straffrundor och liggande"
                f" tid på {liggande_tid} sekunder).\n")

        # Stående skytte
        if kilometer in staende_skytte_index:
            skyttelage = "staende"

            # Rätt ordning vid skidskyttet
            index = stallning.index(skidskytt)
            if index < len(stallning) - 1:
                stallning.insert(index + 1, stallning.pop(index + 1))

            total_min, total_sec = divmod(int(skidskytt["total_tid"]), 60)
            tid_sedan_ledare = ledare_tid - skidskytt["total_tid"]
            tid_sedan_ledare_min = abs(int(tid_sedan_ledare)) // 60
            tid_sedan_ledare_sec = abs(int(tid_sedan_ledare)) % 60
            tid_sedan_ledare_min = -tid_sedan_ledare_min if tid_sedan_ledare < 0 else tid_sedan_ledare_min

            if tid_sedan_ledare == 0:
                print(
                    f"{i + 1}. {skidskytt['namn']}({skidskytt['straffrundor_totalt']}) - {total_min:02d}:"
                    f"{total_sec:02d}, i ledning. Träffsäkerhet: {skidskytt[skyttelage + '_traffsakerhet']}%.")
            else:
                print(
                    f"{i + 1}. {skidskytt['namn']}({skidskytt['straffrundor_totalt']}) - {total_min:02d}:"
                    f"{total_sec:02d}, {abs(tid_sedan_ledare_min)}m {abs(tid_sedan_ledare_sec)}s efter ledaren."
                    f" Träffsäkerhet: {skidskytt[skyttelage + '_traffsakerhet']}%.")

            # Uppdatera träffsäkerheten och utför det liggande skyttet
            ursprunglig_staende_traffsakerhet = skidskytt["staende_traffsakerhet"]
            uppdatera_traffsakerhet(skidskytt, "staende")
            straffrundor = 0
            traffar = 0


            # Check if it's the last shooting in the competition and if the skier is in top 3
            if kilometer == max(staende_skytte_index) and i in range(5):
                time_delay = 3
            else:
                time_delay = 0

            for skott in range(5):
                slump = random.randint(1, 100)
                if slump <= skidskytt[f"{skyttelage}_traffsakerhet"]:
                    traffar += 1
                else:
                    straffrundor += 1
                print(
                    f"Skott {skott + 1}: Träffar: {traffar} av 5 - straffrundor {straffrundor}.")

                if i < 10:
                    time.sleep(0.8)
                elif i > 15:
                    time.sleep(0.1)
                else:
                    time.sleep(0)

            ursprunglig_staende_tid = skidskytt["staende_tid"]
            staende_tid = skidskytt["staende_tid"] + random.randint(-5, 5)
            print(f"Stående tid: {staende_tid} sekunder.")

            if traffar == 5:
                print(f"Fantastiskt! {Fore.GREEN}{skidskytt['namn']} skjuter fullt!{Style.RESET_ALL}")
            elif traffar == 3:
                print(f"Svagt skytte av {skidskytt['namn']}.")
            elif traffar < 3:
                print(f"{Fore.RED}Vilken mardröm för {skidskytt['namn']}!{Style.RESET_ALL}")

            straff_tidstillagg = 20
            tid = staende_tid + (straffrundor * straff_tidstillagg)
            skidskytt.update({
                "staende_tid": ursprunglig_staende_tid,
                "staende_traffar": skidskytt.get("staende_traffar", 0) + traffar,
                "straffrundor": straffrundor,
                "straffrundor_totalt": skidskytt.get("straffrundor_totalt", 0) + straffrundor,
                "staende_traffsakerhet": ursprunglig_staende_traffsakerhet
            })
            print(
                f"Total tid för skyttet: {tid} sekunder ({traffar} träffar, {straffrundor} straffrundor och stående"
                f" tid på {staende_tid} sekunder).\n")

        # Beräkna total tid för varje skidskytt för varje kilometer och totalt
        tidstillagg = skidskytt["tidstillagg"]
        skidskytt[f"kilometer_{kilometer}_total_tid"] = skidskytt.get(f"kilometer_{kilometer-1}_total_tid", 0) + tid + grundtid + tidstillagg
        skidskytt["total_tid"] = skidskytt[f"kilometer_{kilometer}_total_tid"]

    # Sortera skidskyttarna efter total tid för den aktuella kilometer
    stallning = sorted(skidskyttar, key=lambda x: x[f"kilometer_{kilometer}_total_tid"])

    # Skriv ut ställningen för varje kilometer, samt avstånd till ledaren
    for i, skidskytt in enumerate(stallning):
        ledare_tid = stallning[0]["total_tid"]
        tid_sedan_ledare = round(skidskytt["total_tid"] - ledare_tid, 1)
        total_min, total_sec = divmod(int(skidskytt["total_tid"]), 60)
        kilometer_min, kilometer_sec = divmod(int(skidskytt[f"kilometer_{kilometer}_tid"]), 60)

        # Räkna ut avståndet till ledaren, visa det i sekunder om det är mindre än en minut - annars i minut+sek
        if tid_sedan_ledare >= 60:
            tid_sedan_ledare_min, tid_sedan_ledare_sec = divmod(int(tid_sedan_ledare), 60)
            tid_sedan_ledare_str = f"{tid_sedan_ledare_min}m{tid_sedan_ledare_sec}s"
        else:
            tid_sedan_ledare_str = f"{tid_sedan_ledare:.1f}s"

        # Time delay för åkarna som ligger 1-5
        sleep_time = 0.9 if i < 6 else (0.3 if i < 10 else 0)
        time.sleep(sleep_time)

        # Skriv ut ställningen för vare kilometer
        print(
            f"{i + 1:2}. {skidskytt['namn']:<30} ({str(skidskytt['straffrundor_totalt']).rjust(0)}) {total_min: >2d}:"
            f"{total_sec:02d}\t+{tid_sedan_ledare_str:>0}")

    kilometer += 1

# Skriv ut resultatet

print_stallning(sorted(skidskyttar, key=lambda x: (x["straffrundor_totalt"], x["total_tid"])), vald_tavling)
time.sleep(1)
save_excel(skidskyttar, vald_tavling)
main_menu()
# Close the database connection
con.close()