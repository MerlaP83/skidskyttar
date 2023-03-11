import random
import time
import sys,os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import sqlite3

# Connect to the SQLite database
con = sqlite3.connect('skidskytte.db')

# Prompt the user to enter their choice of "damer" or "herrar"
val = input("Välj damer (1) eller herrar (2): ")

# Define the SQL query based on user input
if val == "1":
    gender = "f"
    sql = 'SELECT * FROM skidskyttar WHERE kon = ?'
elif val == "2":
    gender = "m"
    sql = 'SELECT * FROM skidskyttar WHERE kon = ?'
else:
    print("Felaktigt val. Avslutar programmet.")
    sys.exit()

# Hämta tävlingsdata från tabellen "competitions"
con.row_factory = sqlite3.Row   #   add this row
cursor = con.cursor()
# This line is that you need
cursor.execute("SELECT * FROM tavling")
tavlingar = cursor.fetchall()

# Hämta de 40 åkarna med lägst tidstillägg
cursor.execute(f"SELECT * FROM skidskyttar WHERE kon = ? ORDER BY tidstillagg ASC LIMIT 30", (gender,))
skidskyttar = []
weights = [8, 1, 1]  # Viktning för varje alternativ: [0, -1, 1]
for row in cursor:
    skidskytt = dict(row)
    skidskytt["form"] = random.choices([0, -1, 1], weights=weights)[0]
    skidskyttar.append(skidskytt)

# Välj ytterligare 10 slumpmässiga åkare från databasen som inte redan finns i skidskyttar
cursor.execute(sql, (gender,))
remaining_skidskyttar = [dict(row) for row in cursor if dict(row)["id"] not in [sk["id"] for sk in skidskyttar]]
additional_skidskyttar = random.sample(remaining_skidskyttar, 10)

# Lägg till de slumpmässiga åkarna till listan skidskyttar
skidskyttar += additional_skidskyttar


# Grundtid per kilometer i sekunder
grundtid = 130

bra_form = []
dalig_form = []

# Beräkningar för förhandsfavoriter
for skidskytt in skidskyttar:
    fav_tidstillagg = skidskytt["tidstillagg"] + skidskytt["form"]
    fav_form = skidskytt["form"]
    fav_liggande_traffsakerhet = skidskytt["liggande_traffsakerhet"] / 100
    fav_staende_traffsakerhet = skidskytt["staende_traffsakerhet"] / 100
    fav_viktad_poang = (fav_tidstillagg * 3) + 100 / (fav_liggande_traffsakerhet + fav_staende_traffsakerhet) + fav_form
    skidskytt["viktad_poäng"] = fav_viktad_poang + random.randint(-3, 3)

# Sortera förhandsfavoriterna efter viktad poäng och välj de fem bästa
favoriter = sorted(skidskyttar, key=lambda x: x["viktad_poäng"])[:5]

# Funktion för att skriva ut resultatlistan och spara den i databasen
def print_stallning(skidskyttar, tavling):
    print("\nResultat:")
    skidskyttar = sorted(skidskyttar, key=lambda x: x["total_tid"])
    medaljer = ["🥇 ", "🥈 ", "🥉 "]
    with sqlite3.connect('skidskytte.db') as con:
        cursor = con.cursor()
        for i, skidskytt in enumerate(skidskyttar):
            tid_sedan_ledare = round(skidskytt["total_tid"] - skidskyttar[0]["total_tid"], 1)
            tid_sedan_ledare_min, tid_sedan_ledare_sec = divmod(int(tid_sedan_ledare), 60)
            total_min, total_sec = divmod(int(skidskytt["total_tid"]), 60)
            total_tid_str = f"{total_min:02d}:{total_sec:02d}"
            tid_sedan_ledare_str = f"+{tid_sedan_ledare_min}m {tid_sedan_ledare_sec}s" if tid_sedan_ledare > 0 else ""
            if i < 3:
                prefix = medaljer[i]
                print(f"{prefix}{skidskytt['namn']} ({total_tid_str} {tid_sedan_ledare_str})"
                      f" - straffrundor: {skidskytt['straffrundor_totalt']}.")
            else:
                print(f"{i+1}. {skidskytt['namn']} ({total_tid_str} {tid_sedan_ledare_str})"
                      f" - straffrundor: {skidskytt['straffrundor_totalt']}.")
            # Spara resultatet i databasen
            cursor.execute(
                "INSERT INTO resultat (placering, id, namn, total_tid, tid_sedan_ledare, straffrundor_totalt, tavling) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (i + 1, skidskytt["id"], skidskytt["namn"], skidskytt["total_tid"], tid_sedan_ledare_str,
                 skidskytt["straffrundor_totalt"], tavling["namn"]))


# Slumpmässiga händelser

def hantera_handelse(skidskytt, tid, kilometer):
    handelser = {
        1: {"tid": 10, "text": "får problem med ena skidan!"},
        2: {"tid": 15, "text": "ena stav går av!"},
        3: {"tid": 5, "text": "verkar ha dåligt glid!"},
        4: {"tid": 10, "text": "vurpar!"},
        5: {"tid": 5, "text": "tappar en stav och tvingas hämta den."},
        6: {"tid": -5, "text": "får en burst av energi och ökar tempot."},
        7: {"tid": -3, "text": "kämpar på bra."},
        8: {"tid": -5, "text": "kämpar extra hårt efter publikens hejarop."},
        9: {"tid": -3, "text": "åker tekniskt skickligt idag."},
        10: {"tid": -8, "text": "åker överraskande snabbt!"},
        11: {"tid": -5, "text": "har hittat ett bättre glid och ökar tempot."},
        12: {"tid": -5, "text": "ökar takten och ser ut att orka hålla ett högre tempo."},
        13: {"tid": -5, "text": "är stark i motlut och tar in tid på konkurrenterna."},
        14: {"tid": -3, "text": "är en av de snabbaste i nedförsbackarna."},
        15: {"tid": 10, "text": "tappar en skida och tvingas stanna för att sätta fast den igen."},
        16: {"tid": 15, "text": "kraschar in i en annan åkare och tappar tid."},
        17: {"tid": 8, "text": "känner av en muskelskada och tvingas sänka tempot."},
        18: {"tid": 15, "text": "åker fel och måste vända om för att komma tillbaka till rätt bana."},
        19: {"tid": 5, "text": "ser tung ut i åkningen."},
        20: {"tid": 10, "text": "har problem med en skidbindning och tvingas stanna för att åtgärda det."},
        21: {"tid": 15, "text": "ser ut att ha vallat bort sig idag."},
        22: {"tid": -5, "text": "har bra flyt i åkningen."},
        23: {"tid": -5, "text": "får extra energi och orkar åka fortare."},
        24: {"tid": -7, "text": "har utvecklats i tekniken och åker betydligt bättre idag än tidigare."},
        25: {"tid": -8, "text": "har en extra stark dag och kan åka med mycket kraft i åkningen."},
        26: {"tid": -10, "text": "har bättre glid än konkurrenterna!"}
    }

    handelse = random.randint(1, 600)
    if handelse in handelser:
        handelsebeskrivning = handelser.get(handelse)
        tid += handelsebeskrivning["tid"]
        skidskytt[f"kilometer_{kilometer}_handelse"] = handelsebeskrivning["text"]
        print(f"{skidskytt['namn']} {handelsebeskrivning['text']}")
        time.sleep(1)
    # Beräkna total tid för varje skidskytt för varje kilometer och totalt
    skidskytt[f"kilometer_{kilometer}_total_tid"] = (skidskytt.get(f"kilometer_{kilometer-1}_total_tid") or 0)
    skidskytt["total_tid"] = skidskytt[f"kilometer_{kilometer}_total_tid"]
    return tid

# Funktion för att uppdatera träffsäkerheten beroende på vinden
def uppdatera_traffsakerhet(skidskytt, skyttelage):
    skyttelag = skidskytt.get(f"{skyttelage}_traffsakerhet")
    if skyttelag is None:
        print("Felaktigt skyttelage angivet.")
        return

    vindstyrka = vindstyrka1 + random.randint(-2, 2)
    beskrivningar = {-6: "Mycket bra", -5: "Bra", -4: "Bra", -3: "Bra", -2: "Bra", -1: "Bra", 0: "Neutrala", 1: "Svåra",
                     2: "Svåra", 3: "Svåra", 4: "Mycket svåra", 5: "Mycket svåra", 6: "Extremt svåra"}
    beskrivning = beskrivningar.get(vindstyrka, "Ogiltig vindstyrka") + " vindförhållanden."

    print("Aktuell vindstyrka:",beskrivning)

    ny_traffsakerhet = skyttelag - vindstyrka - (skidskytt["form"] * 5)
    skidskytt[f"{skyttelage}_traffsakerhet"] = max(min(ny_traffsakerhet, 100), 0)

def clear_screen():
    os.system("cls" if os.name == "nt" else "clear")

# Låt användaren välja tävling
print("Välj tävling:")
for i, tavling in enumerate(tavlingar):
    print(f"{i+1}. {tavling['namn']} - {tavling['gren']}.")
try:
    vald_tavling_index = int(input("Ange vald tävling: ")) - 1
    vald_tavling = tavlingar[vald_tavling_index]
except (ValueError, IndexError):
    print("Felaktig inmatning, väljer första tävlingen som default.")
    vald_tavling = tavlingar[0]


# Definiera antal kilometer
antal_kilometer = vald_tavling['distans']

# Definiera index för kilometer med liggande skytte
liggande_skytte_index = vald_tavling['liggande_skytte']
if isinstance(liggande_skytte_index, int):
    liggande_skytte_index = [liggande_skytte_index]
else:
    liggande_skytte_index = [int(x) for x in liggande_skytte_index.split(',')]
staende_skytte_index = vald_tavling['staende_skytte']
if isinstance(staende_skytte_index, int):
    staende_skytte_index = [staende_skytte_index]
else:
    staende_skytte_index = [int(x) for x in staende_skytte_index.split(',')]


# Definiera ett högt startvärde för ledande tid
ledande_tid_min = 0
ledande_tid_sec = 0
ledande_skytt = ""

# Slumpa fram vindstyrkan
vindstyrka1 = random.randint(-1, 1)

# Spelet börjar
print(f"\nVälkomna till {vald_tavling['namn']} i {vald_tavling['plats']}.\nI dag blir det en {vald_tavling['gren']} på"
      f" {vald_tavling['distans']} km. Det kommer att bli liggande skytte på kilometer "
      f"{vald_tavling['liggande_skytte']} och stående skytte på kilometer {vald_tavling['staende_skytte']} .\n")

# Skriv ut favoriterna
print("Förhandsfavoriter:")
for i, skidskytt in enumerate(favoriter):
    print(f"{i + 1}. {skidskytt['namn']} (viktad poäng: {skidskytt['viktad_poäng']:.2f}, tidstillägg:"
          f" {skidskytt['tidstillagg']}, liggande träffsäkerhet: {skidskytt['liggande_traffsakerhet']:.1f},"
          f" stående träffsäkerhet: {skidskytt['staende_traffsakerhet']:.1f}, form: {skidskytt['form']})")

# Skriv ut namnen på skidskyttarna med bra form
for skidskytt in skidskyttar:
    if skidskytt["form"] == -1:
        bra_form.append(skidskytt["namn"])
    elif skidskytt["form"] == 1:
        dalig_form.append(skidskytt["namn"])

print(f"\nSkidskyttar som väntas ha en bra dag är: {', '.join(bra_form)}.")
print(f"De som väntas ha en sämre dag är: {', '.join(dalig_form)}.\n")
if vindstyrka1 <= -1:
    print("Bra vindförhållanden.")
elif vindstyrka1 <= 0:
    print("Måttliga vindförhållanden.")
else:
    print("Svåra vindförhållanden.")

time.sleep(4)

for kilometer in range(antal_kilometer):
    # Sortera skidskyttarna efter total tid för den aktuella kilometer
    stallning = sorted(skidskyttar, key=lambda x: x.get(f"kilometer_{kilometer}_total_tid", float('inf')))
    plats = 1
    for skidskytt in stallning:
        skidskytt[f"kilometer_{kilometer}_position"] = plats
        plats += 1
    print(f"\nKilometer {kilometer + 1}:")

    # Sortera skidskyttarna efter total tid över alla kilometer
    stallning = sorted(skidskyttar, key=lambda x: x["total_tid"])
    ledare_tid = stallning[0]["total_tid"]

    for i, skidskytt in enumerate(stallning):
        form = skidskytt["form"]
        tarning = random.randint(1, 6)
        tid = form + tarning

        if kilometer not in liggande_skytte_index and kilometer not in staende_skytte_index:
            tid = hantera_handelse(skidskytt, tid, kilometer)

        skidskytt[f"kilometer_{kilometer}_tid"] = tid

        # Liggande skytte
        if kilometer in liggande_skytte_index:

            # Uppdatera skyttens träffsäkerhet
            skyttelage = "liggande"

            # Rätt ordning vid skidskyttet

            index = stallning.index(skidskytt)
            if index < len(stallning) - 1:
                stallning.insert(index + 1, stallning.pop(index + 1))

            total_min, total_sec = divmod(int(skidskytt["total_tid"]), 60)
            tid_sedan_ledare = ledare_tid - skidskytt["total_tid"]
            tid_sedan_ledare_min = -int(tid_sedan_ledare) // 60
            tid_sedan_ledare_sec = abs(int(tid_sedan_ledare)) % 60

            if tid_sedan_ledare == 0:
                print(
                    f"{i + 1}. {skidskytt['namn']}({skidskytt['straffrundor_totalt']}) - {total_min:02d}:"
                    f"{total_sec:02d}, i ledning. Träffsäkerhet: {skidskytt[skyttelage + '_traffsakerhet']}%.")
            else:
                print(
                    f"{i + 1}. {skidskytt['namn']}({skidskytt['straffrundor_totalt']}) - {total_min:02d}:"
                    f"{total_sec:02d}, {abs(tid_sedan_ledare_min)}m {tid_sedan_ledare_sec}s efter ledaren."
                    f" Träffsäkerhet: {skidskytt[skyttelage + '_traffsakerhet']}%.")

            # Uppdatera träffsäkerheten och utför det liggande skyttet
            ursprunglig_liggande_traffsakerhet = skidskytt["liggande_traffsakerhet"]
            uppdatera_traffsakerhet(skidskytt, "liggande")
            straffrundor = 0
            traffar = 0
            for skott in range(5):
                slump = random.randint(1, 100)
                if slump <= skidskytt[f"{skyttelage}_traffsakerhet"]:
                    traffar += 1
                else:
                    straffrundor += 1
                print(
                    f"Skott {skott + 1}: Träffar: {traffar} av 5 - straffrundor {straffrundor}.")
                # Vänta längre tid vid de första skyttarna för att skapa spänning
                if i < 10:
                    time.sleep(0.8)
                else:
                    time.sleep(0.1)

            ursprunglig_liggande_tid = skidskytt["liggande_tid"]
            liggande_tid = skidskytt["liggande_tid"] + random.randint(-5, 5)
            print(f"Liggande tid: {liggande_tid} sekunder.")

            if traffar == 5:
                print(f"Fantastiskt! {skidskytt['namn']} skjuter fullt!")
            elif traffar == 3:
                print(f"Svagt skytte av {skidskytt['namn']}.")
            elif traffar < 3:
                print(f"Vilken mardröm för {skidskytt['namn']}.")

            straff_tidstillagg = 20
            tid = liggande_tid + (straffrundor * straff_tidstillagg)
            skidskytt.update({
                "liggande_tid": ursprunglig_liggande_tid,
                "liggande_traffar": traffar,
                "straffrundor": straffrundor,
                "straffrundor_totalt": skidskytt.get("straffrundor_totalt", 0) + straffrundor,
                "liggande_traffsakerhet": ursprunglig_liggande_traffsakerhet
            })
            print(
                f"Total tid för skyttet: {tid} sekunder ({traffar} träffar, {straffrundor} straffrundor och liggande"
                f" tid på {liggande_tid} sekunder).\n")

        # Stående skytte
        if kilometer in staende_skytte_index:
            skyttelage = "staende"

            # Rätt ordning vid skidskyttet
            index = stallning.index(skidskytt)
            if index < len(stallning) - 1:
                stallning.insert(index + 1, stallning.pop(index + 1))

            total_min, total_sec = divmod(int(skidskytt["total_tid"]), 60)
            tid_sedan_ledare = ledare_tid - skidskytt["total_tid"]
            tid_sedan_ledare_min = abs(int(tid_sedan_ledare)) // 60
            tid_sedan_ledare_sec = abs(int(tid_sedan_ledare)) % 60
            tid_sedan_ledare_min = -tid_sedan_ledare_min if tid_sedan_ledare < 0 else tid_sedan_ledare_min

            if tid_sedan_ledare == 0:
                print(
                    f"{i + 1}. {skidskytt['namn']}({skidskytt['straffrundor_totalt']}) - {total_min:02d}:"
                    f"{total_sec:02d}, i ledning. Träffsäkerhet: {skidskytt[skyttelage + '_traffsakerhet']}%.")
            else:
                print(
                    f"{i + 1}. {skidskytt['namn']}({skidskytt['straffrundor_totalt']}) - {total_min:02d}:"
                    f"{total_sec:02d}, {abs(tid_sedan_ledare_min)}m {abs(tid_sedan_ledare_sec)}s efter ledaren."
                    f" Träffsäkerhet: {skidskytt[skyttelage + '_traffsakerhet']}%.")

            # Uppdatera träffsäkerheten och utför det liggande skyttet
            ursprunglig_staende_traffsakerhet = skidskytt["staende_traffsakerhet"]
            uppdatera_traffsakerhet(skidskytt, "staende")
            straffrundor = 0
            traffar = 0


            # Check if it's the last shooting in the competition and if the skier is in top 3
            if kilometer == max(staende_skytte_index) and i in range(5):
                time_delay = 3
            else:
                time_delay = 0

            for skott in range(5):
                slump = random.randint(1, 100)
                if slump <= skidskytt[f"{skyttelage}_traffsakerhet"]:
                    traffar += 1
                else:
                    straffrundor += 1
                print(
                    f"Skott {skott + 1}: Träffar: {traffar} av 5 - straffrundor {straffrundor}.")

                if i < 10:
                    time.sleep(0.8)
                else:
                    time.sleep(0.1)
                # Add time delay for top 3 skiers in the last shooting
                if time_delay and skott < 4:
                    time.sleep(0.8)
                else:
                    time.sleep(0.1)

            ursprunglig_staende_tid = skidskytt["staende_tid"]
            staende_tid = skidskytt["staende_tid"] + random.randint(-5, 5)
            print(f"Stående tid: {staende_tid} sekunder.")

            if traffar == 5:
                print(f"Fantastiskt! {skidskytt['namn']} skjuter fullt!")
            elif traffar == 3:
                print(f"Svagt skytte av {skidskytt['namn']}.")
            elif traffar < 3:
                print(f"Vilken mardröm för {skidskytt['namn']}.")

            straff_tidstillagg = 20
            tid = staende_tid + (straffrundor * straff_tidstillagg)
            skidskytt.update({
                "staende_tid": ursprunglig_staende_tid,
                "staende_traffar": traffar,
                "straffrundor": straffrundor,
                "straffrundor_totalt": skidskytt.get("straffrundor_totalt", 0) + straffrundor,
                "staende_traffsakerhet": ursprunglig_staende_traffsakerhet
            })
            print(
                f"Total tid för skyttet: {tid} sekunder ({traffar} träffar, {straffrundor} straffrundor och stående"
                f" tid på {staende_tid} sekunder).\n")

        # Beräkna total tid för varje skidskytt för varje kilometer och totalt
        tidstillagg = skidskytt["tidstillagg"]
        skidskytt[f"kilometer_{kilometer}_total_tid"] = skidskytt.get(f"kilometer_{kilometer-1}_total_tid", 0) + tid + grundtid + tidstillagg
        skidskytt["total_tid"] = skidskytt[f"kilometer_{kilometer}_total_tid"]

    # Sortera skidskyttarna efter total tid för den aktuella kilometer
    stallning = sorted(skidskyttar, key=lambda x: x[f"kilometer_{kilometer}_total_tid"])

    # Skriv ut ställningen för varje kilometer, samt avstånd till ledaren
    for i, skidskytt in enumerate(stallning):
        ledare_tid = stallning[0]["total_tid"]
        tid_sedan_ledare = round(skidskytt["total_tid"] - ledare_tid, 1)
        total_min, total_sec = divmod(int(skidskytt["total_tid"]), 60)
        kilometer_min, kilometer_sec = divmod(int(skidskytt[f"kilometer_{kilometer}_tid"]), 60)

        # Räkna ut avståndet till ledaren, visa det i sekunder om det är mindre än en minut - annars i minut+sek
        if tid_sedan_ledare >= 60:
            tid_sedan_ledare_min, tid_sedan_ledare_sec = divmod(int(tid_sedan_ledare), 60)
            tid_sedan_ledare_str = f"{tid_sedan_ledare_min}m{tid_sedan_ledare_sec}s"
        else:
            tid_sedan_ledare_str = f"{tid_sedan_ledare:.1f}s"

        # Time delay för åkarna som ligger 1-5
        sleep_time = 1 if i < 6 else (0.3 if i < 10 else 0)
        time.sleep(sleep_time)

        # Skriv ut ställningen för vare kilometer
        print(
            f"{i + 1:2}. {skidskytt['namn']:<30} ({str(skidskytt['straffrundor_totalt']).rjust(0)}) {total_min: >2d}:"
            f"{total_sec:02d}\t+{tid_sedan_ledare_str:>0}")

    kilometer += 1

# Skriv ut resultatet

print_stallning(sorted(skidskyttar, key=lambda x: (x["straffrundor_totalt"], x["total_tid"])), vald_tavling)
time.sleep(1)

# Skapa rubriker
headers = [
    "Placering",
    "Namn",
    "Total Tid",
    "Straffrundor",
    "Liggande Träffar",
    "Stående Träffar"
]

# Sortera skidskyttar baserat på totala tiden och lagra placeringen
skidskyttar = sorted(skidskyttar, key=lambda k: k['total_tid'])
for i, skidskytt in enumerate(skidskyttar):
    skidskytt['placering'] = i + 1

# Skapa en lista med data för varje skidskytt
skidskytt_data = [
    {
        "Placering": skidskytt["placering"],
        "Namn": skidskytt["namn"],
        "Total Tid": f'{int(skidskytt["total_tid"]//60)}:{int(skidskytt["total_tid"]%60):02}',
        "Straffrundor": skidskytt["straffrundor_totalt"],
        "Liggande Träffar": skidskytt["liggande_traffar"],
        "Stående Träffar": skidskytt["staende_traffar"]
    }
    for skidskytt in skidskyttar
]

# Skapa ett Excel-dokument
workbook = Workbook()
worksheet = workbook.active

# Skriv rubrikerna till Excel
for i, header in enumerate(headers):
    worksheet.cell(row=1, column=i+1, value=header).font = Font(bold=True)
    worksheet.column_dimensions[get_column_letter(i+1)].width = len(header) + 2

# Skriv data till Excel
for i, data in enumerate(skidskytt_data):
    for j, header in enumerate(headers):
        worksheet.cell(row=i+2, column=j+1, value=data[header]).alignment = Alignment(horizontal="center")

# Spara Excel-dokumentet
workbook.save(vald_tavling['namn'] + ".xlsx")

# Close the database connection
con.close()